from openpyxl import load_workbook
import sqlite3
import mysql.connector

#excel
workbook = load_workbook('รายงานสินค้า.xlsx')
sheet = workbook.active

values = []
#loop ทีละแถว
for row in sheet.iter_rows(min_row = 2,values_only = True):
    print(row)
    values.append(row) 

con = sqlite3.connect('stockdb.db')
cursur = con.cursor()
data = '''
    INSERT INTO products(id,product_id,title,unit,amount)
    VALUES(?,?,?,?,?);
    '''
    
cursur.executemany(data,values)
con.commit()
print('Insert :'+str(cursur.rowcount))